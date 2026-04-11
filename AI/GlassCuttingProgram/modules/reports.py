#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Reports Module
Generate reports and analytics for glass cutting program

Features:
- Daily/Weekly/Monthly production reports
- Utilization analysis
- Waste analysis
- Order status reports
- PDF export (reportlab)
- Excel export (openpyxl)
- Charts and graphs
"""

import os
import io
from datetime import datetime, timedelta
from typing import List, Dict, Optional
from pathlib import Path
from dataclasses import dataclass

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("Warning: reportlab not installed. Run: pip install reportlab")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Run: pip install openpyxl")


@dataclass
class ReportData:
    """Report data structure"""
    report_type: str
    title: str
    period_start: datetime
    period_end: datetime
    generated_at: datetime
    data: Dict
    summary: Dict


class ReportGenerator:
    """Generate various reports"""
    
    def __init__(self, output_dir: Optional[str] = None):
        self.output_dir = Path(output_dir) if output_dir else Path(__file__).parent.parent / 'output' / 'reports'
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Styles
        self.styles = getSampleStyleSheet() if REPORTLAB_AVAILABLE else None
    
    def generate_daily_report(self, date: datetime, data: Dict) -> ReportData:
        """Generate daily production report"""
        report = ReportData(
            report_type='daily',
            title=f'Günlük Üretim Raporu - {date.strftime("%d.%m.%Y")}',
            period_start=date.replace(hour=0, minute=0, second=0),
            period_end=date.replace(hour=23, minute=59, second=59),
            generated_at=datetime.now(),
            data=data,
            summary={
                'total_orders': data.get('total_orders', 0),
                'completed_orders': data.get('completed_orders', 0),
                'total_cuts': data.get('total_cuts', 0),
                'avg_utilization': data.get('avg_utilization', 0),
                'total_waste': data.get('total_waste', 0),
                'total_time': data.get('total_time', 0)
            }
        )
        
        return report
    
    def generate_weekly_report(self, start_date: datetime, data: Dict) -> ReportData:
        """Generate weekly production report"""
        end_date = start_date + timedelta(days=6)
        
        report = ReportData(
            report_type='weekly',
            title=f'Haftalık Üretim Raporu - {start_date.strftime("%d.%m")} - {end_date.strftime("%d.%m.%Y")}',
            period_start=start_date,
            period_end=end_date,
            generated_at=datetime.now(),
            data=data,
            summary={
                'total_orders': data.get('total_orders', 0),
                'daily_average': data.get('daily_average', 0),
                'best_day': data.get('best_day', ''),
                'avg_utilization': data.get('avg_utilization', 0),
                'total_waste': data.get('total_waste', 0)
            }
        )
        
        return report
    
    def generate_monthly_report(self, year: int, month: int, data: Dict) -> ReportData:
        """Generate monthly production report"""
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1) - timedelta(days=1)
        
        report = ReportData(
            report_type='monthly',
            title=f'Aylık Üretim Raporu - {start_date.strftime("%B %Y")}',
            period_start=start_date,
            period_end=end_date,
            generated_at=datetime.now(),
            data=data,
            summary={
                'total_orders': data.get('total_orders', 0),
                'total_sheets': data.get('total_sheets', 0),
                'avg_utilization': data.get('avg_utilization', 0),
                'total_waste_area': data.get('total_waste_area', 0),
                'total_cutting_time': data.get('total_cutting_time', 0),
                'efficiency': data.get('efficiency', 0)
            }
        )
        
        return report
    
    def generate_waste_analysis(self, period: str, data: Dict) -> ReportData:
        """Generate waste analysis report"""
        report = ReportData(
            report_type='waste_analysis',
            title=f'Fire Analizi Raporu - {period}',
            period_start=datetime.now() - timedelta(days=30),
            period_end=datetime.now(),
            generated_at=datetime.now(),
            data=data,
            summary={
                'total_waste': data.get('total_waste', 0),
                'waste_by_type': data.get('waste_by_type', {}),
                'waste_trend': data.get('waste_trend', []),
                'recommendations': data.get('recommendations', [])
            }
        )
        
        return report
    
    def export_to_pdf(self, report: ReportData, filename: Optional[str] = None) -> str:
        """Export report to PDF"""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab not installed")
        
        if not filename:
            filename = f"report_{report.report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        filepath = self.output_dir / filename
        
        # Create PDF document
        doc = SimpleDocTemplate(
            str(filepath),
            pagesize=landscape(A4),
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        # Build content
        content = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1e293b'),
            alignment=TA_CENTER
        )
        content.append(Paragraph(report.title, title_style))
        content.append(Spacer(1, 0.2*inch))
        
        # Summary table
        summary_data = [['Metric', 'Value']]
        for key, value in report.summary.items():
            formatted_key = key.replace('_', ' ').title()
            if isinstance(value, float):
                formatted_value = f"{value:.2f}"
            else:
                formatted_value = str(value)
            summary_data.append([formatted_key, formatted_value])
        
        summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f1f5f9')),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1e293b')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1'))
        ]))
        
        content.append(summary_table)
        content.append(Spacer(1, 0.3*inch))
        
        # Detailed data
        if report.data.get('orders'):
            content.append(Paragraph("Sipariş Detayları", self.styles['Heading2']))
            content.append(Spacer(1, 0.1*inch))
            
            orders_data = [['Order ID', 'Glass Type', 'Size', 'Utilization', 'Status']]
            for order in report.data['orders'][:20]:  # Limit to 20
                orders_data.append([
                    order.get('order_id', ''),
                    order.get('glass_type', ''),
                    f"{order.get('width', 0)}x{order.get('height', 0)}",
                    f"{order.get('utilization', 0)*100:.1f}%",
                    order.get('status', '')
                ])
            
            orders_table = Table(orders_data, colWidths=[1.2*inch, 1*inch, 1.2*inch, 1*inch, 1*inch])
            orders_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f8fafc'), colors.white])
            ]))
            
            content.append(orders_table)
        
        # Build PDF
        doc.build(content)
        
        return str(filepath)
    
    def export_to_excel(self, report: ReportData, filename: Optional[str] = None) -> str:
        """Export report to Excel"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl not installed")
        
        if not filename:
            filename = f"report_{report.report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        filepath = self.output_dir / filename
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = report.report_type
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Title
        ws['A1'] = report.title
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:E1')
        
        # Summary section
        ws['A3'] = "Summary Metrics"
        ws['A3'].font = Font(bold=True, size=14)
        
        row = 4
        for key, value in report.summary.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = value
            row += 1
        
        # Orders data
        if report.data.get('orders'):
            ws['G3'] = "Order Details"
            ws['G3'].font = Font(bold=True, size=14)
            
            # Headers
            headers = ['Order ID', 'Glass Type', 'Width', 'Height', 'Utilization', 'Status']
            for col, header in enumerate(headers, 7):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Data
            for i, order in enumerate(report.data['orders'], 5):
                ws.cell(row=i, column=7).value = order.get('order_id', '')
                ws.cell(row=i, column=8).value = order.get('glass_type', '')
                ws.cell(row=i, column=9).value = order.get('width', 0)
                ws.cell(row=i, column=10).value = order.get('height', 0)
                ws.cell(row=i, column=11).value = order.get('utilization', 0)
                ws.cell(row=i, column=12).value = order.get('status', '')
        
        # Add chart if data available
        if report.data.get('utilization_trend'):
            chart_ws = wb.create_sheet(title='Charts')
            
            # Utilization chart
            chart = LineChart()
            chart.title = "Utilization Trend"
            chart.style = 13
            chart.x_axis.title = 'Date'
            chart.y_axis.title = 'Utilization %'
            
            # Add data to chart (simplified)
            # In real implementation, add actual data
            
            chart_ws.add_chart(chart, "A1")
        
        # Save workbook
        wb.save(str(filepath))
        
        return str(filepath)
    
    def generate_json_report(self, report: ReportData, filename: Optional[str] = None) -> str:
        """Export report to JSON"""
        import json
        
        if not filename:
            filename = f"report_{report.report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        filepath = self.output_dir / filename
        
        report_data = {
            'report_type': report.report_type,
            'title': report.title,
            'period': {
                'start': report.period_start.isoformat(),
                'end': report.period_end.isoformat()
            },
            'generated_at': report.generated_at.isoformat(),
            'summary': report.summary,
            'data': report.data
        }
        
        with open(filepath, 'w') as f:
            json.dump(report_data, f, indent=2)
        
        return str(filepath)


class AnalyticsEngine:
    """Analyze cutting data for reports"""
    
    def __init__(self):
        pass
    
    def calculate_utilization_stats(self, cutting_history: List[Dict]) -> Dict:
        """Calculate utilization statistics"""
        if not cutting_history:
            return {}
        
        utilizations = [c.get('utilization', 0) for c in cutting_history if c.get('utilization')]
        
        return {
            'avg_utilization': sum(utilizations) / len(utilizations) if utilizations else 0,
            'max_utilization': max(utilizations) if utilizations else 0,
            'min_utilization': min(utilizations) if utilizations else 0,
            'total_cuts': len(cutting_history)
        }
    
    def calculate_waste_stats(self, cutting_history: List[Dict]) -> Dict:
        """Calculate waste statistics"""
        if not cutting_history:
            return {}
        
        total_waste = sum(c.get('waste_area', 0) for c in cutting_history)
        
        return {
            'total_waste_area': total_waste,
            'total_waste_m2': total_waste / 1000000,
            'avg_waste_per_cut': total_waste / len(cutting_history) if cutting_history else 0
        }
    
    def calculate_efficiency(self, cutting_history: List[Dict]) -> Dict:
        """Calculate efficiency metrics"""
        if not cutting_history:
            return {}
        
        total_time = sum(c.get('cutting_time', 0) for c in cutting_history)
        total_cuts = sum(c.get('total_cuts', 0) for c in cutting_history)
        
        return {
            'total_cutting_time': total_time,
            'avg_time_per_cut': total_time / len(cutting_history) if cutting_history else 0,
            'cuts_per_hour': (total_cuts / total_time) * 60 if total_time > 0 else 0
        }
    
    def get_daily_breakdown(self, cutting_history: List[Dict]) -> List[Dict]:
        """Get daily breakdown of production"""
        daily_data = {}
        
        for record in cutting_history:
            date = record.get('created_at', '')[:10]  # YYYY-MM-DD
            if date not in daily_data:
                daily_data[date] = {
                    'date': date,
                    'orders': 0,
                    'utilization': 0,
                    'waste': 0
                }
            
            daily_data[date]['orders'] += 1
            daily_data[date]['utilization'] += record.get('utilization', 0)
            daily_data[date]['waste'] += record.get('waste_area', 0)
        
        # Average utilization per day
        for date in daily_data:
            if daily_data[date]['orders'] > 0:
                daily_data[date]['avg_utilization'] = daily_data[date]['utilization'] / daily_data[date]['orders']
        
        return sorted(daily_data.values(), key=lambda x: x['date'])


def demo():
    """Demo usage"""
    print("=" * 60)
    print("Reports Module Demo")
    print("=" * 60)
    
    # Create generator
    generator = ReportGenerator()
    
    # Sample data
    sample_data = {
        'total_orders': 25,
        'completed_orders': 23,
        'total_cuts': 150,
        'avg_utilization': 0.87,
        'total_waste': 2500000,
        'total_time': 480,
        'orders': [
            {'order_id': 'ORD-001', 'glass_type': 'float', 'width': 500, 'height': 400, 'utilization': 0.92, 'status': 'completed'},
            {'order_id': 'ORD-002', 'glass_type': 'laminated', 'width': 600, 'height': 400, 'utilization': 0.85, 'status': 'completed'},
        ]
    }
    
    # Generate daily report
    print("\n📊 Generating daily report...")
    report = generator.generate_daily_report(datetime.now(), sample_data)
    print(f"   Title: {report.title}")
    print(f"   Summary: {report.summary}")
    
    # Export to JSON
    print("\n💾 Exporting to JSON...")
    json_file = generator.generate_json_report(report)
    print(f"   Saved: {json_file}")
    
    # Export to PDF (if available)
    if REPORTLAB_AVAILABLE:
        print("\n📄 Exporting to PDF...")
        pdf_file = generator.export_to_pdf(report)
        print(f"   Saved: {pdf_file}")
    else:
        print("\n⚠️ PDF export skipped (reportlab not installed)")
    
    # Export to Excel (if available)
    if OPENPYXL_AVAILABLE:
        print("\n📊 Exporting to Excel...")
        excel_file = generator.export_to_excel(report)
        print(f"   Saved: {excel_file}")
    else:
        print("\n⚠️ Excel export skipped (openpyxl not installed)")
    
    # Analytics
    print("\n📈 Running analytics...")
    analytics = AnalyticsEngine()
    
    cutting_history = [
        {'utilization': 0.92, 'waste_area': 100000, 'cutting_time': 30, 'total_cuts': 10},
        {'utilization': 0.85, 'waste_area': 150000, 'cutting_time': 45, 'total_cuts': 15},
    ]
    
    util_stats = analytics.calculate_utilization_stats(cutting_history)
    print(f"   Utilization stats: {util_stats}")
    
    waste_stats = analytics.calculate_waste_stats(cutting_history)
    print(f"   Waste stats: {waste_stats}")


if __name__ == "__main__":
    demo()